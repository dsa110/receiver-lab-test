# -*- coding: utf-8 -*-
"""
Created on Fri Jul 26 01:30:39 2019

@author: hidir
"""

###############################################################################
# LAbjack Analog Input Readout ( Single Ended Analog Inputs)
###############################################################################

import u3 , signal , datetime
import numpy as np
import time
from openpyxl import Workbook
import visa as v
import math
#import matplotlib.pyplot as plt
#from openpyxl.chart import (
#    ScatterChart,
#    Reference,
#    Serie,
#)
###############################################################################
# User Settings
###############################################################################
#LabJack
n_avg = 10;
Voffset = -0.009248
# Time Delay in [S] between measurements
delay =  0.1
# Number of samples saved at a time:
z = 1
#file
ftsize = 12


# Read Date and Time from PC clock
# x = datetime.datetime.now()
date_time_string= time.strftime('%m%d%Y %H:%M:%S')
# Format Time 
t = datetime.datetime.strptime(date_time_string,"%m%d%Y %H:%M:%S")


datafilename = str('FEB_Test_Score_Sheet')
folder = 'C:\\Users\\amplab\\Desktop\\FEB_BEB_test\\'
###############################################################################
# Configure Instrument
###############################################################################
#LabJack
print (' Configuring LabJack ')
lj = u3.U3()

lj.configAnalog(0)
lj.configAnalog(1)
lj.configAnalog(2)
lj.configAnalog(3)
lj.configAnalog(4)
lj.configAnalog(5)
lj.configAnalog(6)
lj.configAnalog(7)

lj.configAnalog(8)
lj.configAnalog(9)
lj.configAnalog(10)
lj.configAnalog(11)
lj.configAnalog(12)
lj.configAnalog(13)
lj.configAnalog(14)
lj.configAnalog(15)

#Siglent
print (' Connecting to Instrument ... ')
rm = v.ResourceManager()
rm.list_resources()
spectrum_analyzer = rm.open_resource('USB0::0xF4EC::0x1300::SSA3XLBC1R0061::INSTR')
print(spectrum_analyzer.query("*IDN?"))

spectrum_analyzer.read_termination = '\n'
spectrum_analyzer.query('*OPC?')
time.sleep(0.1)
spectrum_analyzer.query(':SYSTem:TIME?')
time.sleep(0.1)
spectrum_analyzer.query(':SYSTem:DATE?')
time.sleep(0.1)
spectrum_analyzer.write(':DISPlay:WINDow:TRACe:Y:RLEVel -36 DBM')
time.sleep(0.1)
spectrum_analyzer.write(':POWer:ATTenuation 10')
time.sleep(0.1)
spectrum_analyzer.write(':POWer:GAIN OFF') #preamp
time.sleep(0.1)
spectrum_analyzer.write(':UNIT:POWer DBM')
time.sleep(0.1)
spectrum_analyzer.write(':DISPlay:WINDow:TRACe:Y:SPACing LOGarithmic')
time.sleep(0.1)
spectrum_analyzer.write(':DISPlay:WINDow:TRACe:Y:PDIVision 1 dB')
time.sleep(0.1)
spectrum_analyzer.write(':SENSe:CORRection:OFF')
time.sleep(0.1)
#spectrum_analyzer.write(':BWID:AUTO On') # resolution BW
#time.sleep(0.1)

spectrum_analyzer.write(':BWID:AUTO Off') # resolution BW
time.sleep(0.1)
spectrum_analyzer.write(':BWID: 1 MHz') # resolution BW
time.sleep(0.1)

#spectrum_analyzer.write(':BWIDth:VIDeo 1 MHz')
spectrum_analyzer.write(':BWIDth:VIDeo 10 KHz')
time.sleep(0.1)
spectrum_analyzer.write(':TRAC1:MODE WRITE')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARK1:STATe ON')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARK2:STATe ON')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARK3:STATe ON')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARKer1:X 0.36 GHz')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARKer2:X 0.375 GHz')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARKer3:X 0.39 GHz')
time.sleep(0.1)
#spectrum_analyzer.write(':TRAC1:MODE AVERAGE')
#time.sleep(0.1)
spectrum_analyzer.write(':DETector:TRAC1 AVERage')
time.sleep(0.1)
spectrum_analyzer.write(':AVERage:TRACe1:COUNt 16')
time.sleep(0.1)
spectrum_analyzer.write(':CALCulate:MARKer:TABLe ON')
time.sleep(0.1)
spectrum_analyzer.write(':SWEep:MODE AUTO')
time.sleep(0.1)
spectrum_analyzer.write(':SWEep:TIME:AUTO ON')
time.sleep(0.1)
spectrum_analyzer.write(':SWEep:SPEed ACCUracy')
time.sleep(0.1)

# Save data in XLSX  file: 
#Create a new workbook
wb = Workbook()
#create new worksheets 
ws1 = wb.worksheets[0]

# TEMPLATE for Sandy
ws1['A1'] = 'Test Data for FEB Score Sheet'
ws1.merge_cells('A1:D1')
#ws1['E1'] = 'Date & Time:'

# Read Date and Time from PC clock
# x = datetime.datetime.now()
date_time_string= time.strftime('%m%d%Y %H:%M:%S')
# Format Time 
t = datetime.datetime.strptime(date_time_string,"%m%d%Y %H:%M:%S")

#ws1['G1'] = str(t)
#ws1.merge_cells('G1:I1')

ws1['A2'] = 'FEB_SN'
ws1['B2'] = 'BEB_SN'
ws1['C2'] = 'BEB Out, 375 MHz, NG OFF'
ws1['D2'] = 'BEB Out, 375 MHz, NG ON'
ws1['E2'] = 'FEB Y Factor  dB'
ws1['F2'] = 'FEB Y Factor  Ratio'
ws1['G2'] = 'FEB/BEB Tn, K'
ws1['H2'] = 'FEB/BEB NF dB'
ws1['I2'] = 'Tsys Contrib K'
ws1['J2'] = 'BEB Out, w. LNA, 300K in,  dBm/MHz'
ws1['K2'] = 'BEB Out, w. LNA, 300K in,  Total dBm'
ws1['L2'] = 'BEB Out, w. LNA, Tsys=26K,  Total dBm'
ws1['M2'] = 'Test Date'
ws1['N2'] = 'By'
ws1['O2'] = 'FEB Temp  C'
ws1['P2'] = 'BEB PD mA'
ws1['Q2'] = 'FEB mA'
ws1['R2'] = 'FEB dBm mV OFF'
ws1['S2'] = 'FEB dBm mV ON'
ws1['T2'] = 'BEB dBm mV  OFF'
ws1['U2'] = 'BEB dBm mV  ON'
ws1['V2'] = 'FEB LD V'


#Siglent
## Acquire Data
fstart=125
fstop=625
sweep_count=1
spectrum_analyzer.write('*WAI')
time.sleep(0.1)
spectrum_analyzer.write('SENSE:FREQuency:STARt '+str(fstart)+' MHz')
time.sleep(0.1)
spectrum_analyzer.write('SENSE:FREQuency:STOP '+str(fstop)+' MHz')
time.sleep(0.1)
spectrum_analyzer.write(':SWEep:COUNt '+str(sweep_count))
#time.sleep(0.1)
freqpoint=16 # total number of freq points to be averaged when calculating SA data

###############################################################################
# Interrupt handler 
###############################################################################
class GracefulInterruptHandler(object):

    def __init__(self, sig=signal.SIGINT):
        self.sig = sig

    def __enter__(self):

        self.interrupted = False
        self.released = False

        self.original_handler = signal.getsignal(self.sig)

        def handler(signum, frame):
            self.release()
            self.interrupted = True

        signal.signal(self.sig, handler)

        return self

    def __exit__(self, type, value, tb):
        self.release()

    def release(self):

        if self.released:
            return False

        signal.signal(self.sig, self.original_handler)
        
        self.released = True

row=3
print (' Please make sure that the switches are turned towards LabJack on the test box')
print (' Please keep the same BEB during FEB tests')
Tester= input("MOVE CURSOR; Testor Inititals? ")
filename=input("filename:")
BEB_SN = input("Please enter BEB SN :")
while(True):
    if ((BEB_SN[len(BEB_SN)-1])==('A')) or ((BEB_SN[len(BEB_SN)-1])==('B')):
        break
    else:
        BEB_SN = input("Please enter BEB SN with correct Channel (ie. 27A) :")

while(True):
    FEB_SN = input("Please enter FEB SN :")
    ###############################################################################
    # Initialization Save Data 
    ###############################################################################
    #LabJack
    a0_averages_NGENoff = []
    a1_averages_NGENoff = []
    a2_averages_NGENoff = []
    a3_averages_NGENoff = []
    a4_averages_NGENoff = []
    a5_averages_NGENoff = []
    a6_averages_NGENoff = []
    a7_averages_NGENoff = []
    a8_averages_NGENoff = []
    a9_averages_NGENoff = []
    a10_averages_NGENoff = []
    a11_averages_NGENoff = []
    a12_averages_NGENoff = []
    a13_averages_NGENoff = []
    a14_averages_NGENoff = []
    a15_averages_NGENoff = []
    
    a0_averages_NGENon = []
    a1_averages_NGENon = []
    a2_averages_NGENon = []
    a3_averages_NGENon = []
    a4_averages_NGENon = []
    a5_averages_NGENon = []
    a6_averages_NGENon = []
    a7_averages_NGENon = []
    a8_averages_NGENon = []
    a9_averages_NGENon = []
    a10_averages_NGENon = []
    a11_averages_NGENon = []
    a12_averages_NGENon = []
    a13_averages_NGENon = []
    a14_averages_NGENon = []
    a15_averages_NGENon = []
    save_values = []
    
    times = []    
    ################################################################################
    ## Measurement Loop
    ################################################################################
    #LabJack and SA
    dig_offset=0.024
    DAC1_VALUE = lj.voltageToDACBits(3+dig_offset, dacNumber = 1, is16Bits = False)#CAL voltage value is set to 3V
    lj.getFeedback(u3.DAC1_8(DAC1_VALUE))
    
    # Get data when NGEN is OFF
    
    DAC0_VALUE = lj.voltageToDACBits(0-dig_offset+0.010, dacNumber = 1, is16Bits = False)#NGEN is OFF
    lj.getFeedback(u3.DAC0_8(DAC0_VALUE))   
    k = 0
    with GracefulInterruptHandler() as h:
        while True:
            for i in range(n_avg):
                a0 = lj.getAIN(0)
                a1 = lj.getAIN(1)
                a2 = lj.getAIN(2)
                a3 = lj.getAIN(3)
                a4 = lj.getAIN(4)
                a5 = lj.getAIN(5)
                a6 = lj.getAIN(6)
                a7 = lj.getAIN(7)
                a8 = lj.getAIN(8)
                a9 = lj.getAIN(9)
                a10 = lj.getAIN(10)
                a11 = lj.getAIN(11)
                a12 = lj.getAIN(12)
                a13 = lj.getAIN(13)
                a14 = lj.getAIN(14)
                a15 = lj.getAIN(15)
                a0_averages_NGENoff.append(a0)
                a1_averages_NGENoff.append(a1)
                a2_averages_NGENoff.append(a2)
                a3_averages_NGENoff.append(a3)
                a4_averages_NGENoff.append(a4)
                a5_averages_NGENoff.append(a5)
                a6_averages_NGENoff.append(a6)
                a7_averages_NGENoff.append(a7)
                a8_averages_NGENoff.append(a8)
                a9_averages_NGENoff.append(a9)
                a10_averages_NGENoff.append(a10)
                a11_averages_NGENoff.append(a11)
                a12_averages_NGENoff.append(a12)
                a13_averages_NGENoff.append(a13)
                a14_averages_NGENoff.append(a14)
                a15_averages_NGENoff.append(a15)
    
            a0_avg_value_NGENoff = np.average(a0_averages_NGENoff) - Voffset
            a1_avg_value_NGENoff = np.average(a1_averages_NGENoff) - Voffset
            a2_avg_value_NGENoff = np.average(a2_averages_NGENoff) - Voffset
            a3_avg_value_NGENoff = np.average(a3_averages_NGENoff) - Voffset
            a4_avg_value_NGENoff = np.average(a4_averages_NGENoff) - Voffset
            a5_avg_value_NGENoff = np.average(a5_averages_NGENoff) - Voffset
            a6_avg_value_NGENoff = np.average(a6_averages_NGENoff) - Voffset
            a7_avg_value_NGENoff = np.average(a7_averages_NGENoff) - Voffset
            a8_avg_value_NGENoff = np.average(a8_averages_NGENoff) - Voffset
            a9_avg_value_NGENoff = np.average(a9_averages_NGENoff) - Voffset
            a10_avg_value_NGENoff = np.average(a10_averages_NGENoff) - Voffset
            a11_avg_value_NGENoff = np.average(a11_averages_NGENoff) - Voffset
            a12_avg_value_NGENoff = np.average(a12_averages_NGENoff) - Voffset
            a13_avg_value_NGENoff = np.average(a13_averages_NGENoff) - Voffset
            a14_avg_value_NGENoff = np.average(a14_averages_NGENoff) - Voffset
            a15_avg_value_NGENoff = np.average(a15_averages_NGENoff) - Voffset
            print (" NGENoff LJ data is saved ")
            break
            if np.remainder(k,z) == 0 and k != 0:
            #if np.remainder(k,z-1) == 0 and k <> 0:
                print (" I saved " + str(z) + "  values for You! ")
                break
                k = -1
            avg_value = []
            averages = []
            time.sleep(delay)
            k= k + 1
            if h.interrupted:
                print (" Exiting Gracefully ...........") 
                #save_data(times,save_values)
                #datafilename.close()
                lj.close()
                break
    #Siglent
    time.sleep(10) #comment this line if using 100kHz video BW
    # Download LgPwr Data    
    LgPwr_off = []
    lgpwr_off = []
    spectrum_analyzer.write('*WAI')
    time.sleep(0.5)
    lgpwr_off = spectrum_analyzer.query(':TRACe:DATA? 1') #This query command returns the current displayed data
    time.sleep(0.5)
    spectrum_analyzer.write('*WAI')
    time.sleep(0.5)
    print (" NGENoff SA data is saved ")
    lgpwr_off  = lgpwr_off.rsplit(',')
    #lgpwr = lgpwr.replace("\x00\n", "")
    nfreq=len(lgpwr_off)-1
    freq = np.linspace(fstart,fstop,nfreq)
    
    for i in range(int(nfreq)):
        LgPwr_off.append(np.float(lgpwr_off[i]))
        
    #SA_data_NGENoff = np.float(lgpwr_off[int((len(lgpwr_off)/2)-1)])
        
    SA_data_NGENoff_mat=[]
    favstart=int(len(lgpwr_off)/2)-1-round(freqpoint/2) # average start frequency
    for i in range(freqpoint):
        SA_data_NGENoff_mat.append(np.float(lgpwr_off[favstart+i]))  
    SA_data_NGENoff=np.average(SA_data_NGENoff_mat) # averaged NGEN off SA power data
    
    # Get data when NGEN is ON
    
    DAC0_VALUE = lj.voltageToDACBits(1-dig_offset+0.010, dacNumber = 1, is16Bits = False)#NGEN is ON
    lj.getFeedback(u3.DAC0_8(DAC0_VALUE))
    
    k = 0
    with GracefulInterruptHandler() as h:
        while True:
            for i in range(n_avg):
                a0 = lj.getAIN(0)
                a1 = lj.getAIN(1)
                a2 = lj.getAIN(2)
                a3 = lj.getAIN(3)
                a4 = lj.getAIN(4)
                a5 = lj.getAIN(5)
                a6 = lj.getAIN(6)
                a7 = lj.getAIN(7)
                a8 = lj.getAIN(8)
                a9 = lj.getAIN(9)
                a10 = lj.getAIN(10)
                a11 = lj.getAIN(11)
                a12 = lj.getAIN(12)
                a13 = lj.getAIN(13)
                a14 = lj.getAIN(14)
                a15 = lj.getAIN(15)
                a0_averages_NGENon.append(a0)
                a1_averages_NGENon.append(a1)
                a2_averages_NGENon.append(a2)
                a3_averages_NGENon.append(a3)
                a4_averages_NGENon.append(a4)
                a5_averages_NGENon.append(a5)
                a6_averages_NGENon.append(a6)
                a7_averages_NGENon.append(a7)
                a8_averages_NGENon.append(a8)
                a9_averages_NGENon.append(a9)
                a10_averages_NGENon.append(a10)
                a11_averages_NGENon.append(a11)
                a12_averages_NGENon.append(a12)
                a13_averages_NGENon.append(a13)
                a14_averages_NGENon.append(a14)
                a15_averages_NGENon.append(a15)
    
            a0_avg_value_NGENon = np.average(a0_averages_NGENon) - Voffset
            a1_avg_value_NGENon = np.average(a1_averages_NGENon) - Voffset
            a2_avg_value_NGENon = np.average(a2_averages_NGENon) - Voffset
            a3_avg_value_NGENon = np.average(a3_averages_NGENon) - Voffset
            a4_avg_value_NGENon = np.average(a4_averages_NGENon) - Voffset
            a5_avg_value_NGENon = np.average(a5_averages_NGENon) - Voffset
            a6_avg_value_NGENon = np.average(a6_averages_NGENon) - Voffset
            a7_avg_value_NGENon = np.average(a7_averages_NGENon) - Voffset
            a8_avg_value_NGENon = np.average(a8_averages_NGENon) - Voffset
            a9_avg_value_NGENon = np.average(a9_averages_NGENon) - Voffset
            a10_avg_value_NGENon = np.average(a10_averages_NGENon) - Voffset
            a11_avg_value_NGENon = np.average(a11_averages_NGENon) - Voffset
            a12_avg_value_NGENon = np.average(a12_averages_NGENon) - Voffset
            a13_avg_value_NGENon = np.average(a13_averages_NGENon) - Voffset
            a14_avg_value_NGENon = np.average(a14_averages_NGENon) - Voffset
            a15_avg_value_NGENon = np.average(a15_averages_NGENon) - Voffset
            print (" NGENon LJ data is saved ")
            break
            if np.remainder(k,z) == 0 and k != 0:
            #if np.remainder(k,z-1) == 0 and k <> 0:
                print (" I saved " + str(z) + "  values for You! ")
                break
                k = -1
            avg_value = []
            averages = []
            time.sleep(delay)
            k= k + 1
            if h.interrupted:
                print (" Exiting Gracefully ...........") 
                #save_data(times,save_values)
                #datafilename.close()
                lj.close()
                break
    #Siglent
    #spectrum_analyzer.write(':TRAC1:MODE WRITE')
    #time.sleep(0.1)
    time.sleep(10) #comment this line if using 100kHz video BW
    # Download LgPwr Data    
    LgPwr_on = []
    lgpwr_on = []
    spectrum_analyzer.write('*WAI')
    time.sleep(0.5)
    lgpwr_on = spectrum_analyzer.query(':TRACe:DATA? 1') #This query command returns the current displayed data
    time.sleep(0.5)
    spectrum_analyzer.write('*WAI')
    time.sleep(0.5)
    print (" NGENon SA data is saved ")
    lgpwr_on  = lgpwr_on.rsplit(',')
    #lgpwr = lgpwr.replace("\x00\n", "")
    nfreq=len(lgpwr_on)-1
    freq = np.linspace(fstart,fstop,nfreq)
    for i in range(int(nfreq)):
        LgPwr_on.append(np.float(lgpwr_on[i]))
    
    #SA_data_NGENon = np.float(lgpwr_on[int((len(lgpwr_on)/2)-1)])
    SA_data_NGENon_mat=[]
    favstart=int(len(lgpwr_on)/2)-1-round(freqpoint/2) # average start frequency
    for i in range(freqpoint):
        SA_data_NGENon_mat.append(np.float(lgpwr_on[favstart+i]))  
    SA_data_NGENon=np.average(SA_data_NGENon_mat) # averaged NGEN on SA power data
    
    ### Write to Excel
    
    Y_dB=SA_data_NGENon-SA_data_NGENoff
    Y_ratio=10**(((SA_data_NGENon-SA_data_NGENoff)/10))
    NoiseSource_ENR=9.81
    Tcold=300
    Thot=Tcold+290*10**(NoiseSource_ENR/10)
    Tn=(Thot-Tcold*Y_ratio)/(Y_ratio-1)
    # print (" Tn =  ",Tn)
    FEB_BEB_NF_dB=10*math.log10(1+Tn/290)
    BEB_out_dBm_per_MHz=SA_data_NGENoff+35-10*math.log10(1+Tn/290)
    BEB_out_300k_totdBm=BEB_out_dBm_per_MHz+26
    BEB_out_26k_totdBm=BEB_out_300k_totdBm-10.6
    
    if ((BEB_SN[len(BEB_SN)-1])==('A')):
       BEB_PD_mA=2000*a15_avg_value_NGENoff/1000 
       BEB_IF_MON_NGENoff=2000*a12_avg_value_NGENoff
       BEB_IF_MON_NGENon=2000*a12_avg_value_NGENon

    if ((BEB_SN[len(BEB_SN)-1])==('B')):
       BEB_PD_mA=2000*a11_avg_value_NGENoff/1000 
       BEB_IF_MON_NGENoff=2000*a8_avg_value_NGENoff
       BEB_IF_MON_NGENon=2000*a8_avg_value_NGENon

    FEB_temp=(2000*a3_avg_value_NGENoff-500)/10   
    FEB_mA=2000*a2_avg_value_NGENoff
    FEB_IF_MON_NGENoff=2000*a1_avg_value_NGENoff/10
    FEB_IF_MON_NGENon=2000*a1_avg_value_NGENon/10 
    FEB_LD_MON_NGENoff=2000*a0_avg_value_NGENoff/1000
    
    print (" NF is : ",FEB_BEB_NF_dB,"dB")
    print (" BEB PD current: ",BEB_PD_mA,"mA")
    
       
    ws1['A'+str(row)] = FEB_SN
    ws1['B'+str(row)] = BEB_SN
    ws1['C'+str(row)] = float(SA_data_NGENoff)
    ws1['D'+str(row)] = float(SA_data_NGENon)
    ws1['E'+str(row)] = float(Y_dB)
    ws1['F'+str(row)] = float(Y_ratio)
    ws1['G'+str(row)] = float(Tn)
    ws1['H'+str(row)] = float(FEB_BEB_NF_dB)    
    ws1['I'+str(row)] = float(Tn/3162)
    ws1['J'+str(row)] = float(BEB_out_dBm_per_MHz)
    ws1['K'+str(row)] = float(BEB_out_300k_totdBm)
    ws1['L'+str(row)] = float(BEB_out_26k_totdBm)
    ws1['M'+str(row)] = str(t)
    ws1['N'+str(row)] = Tester
    ws1['O'+str(row)] = float(FEB_temp)
    ws1['P'+str(row)] = float(BEB_PD_mA) # BEB channel A
    ws1['Q'+str(row)] = float(FEB_mA)    
    ws1['R'+str(row)] = float(FEB_IF_MON_NGENoff)
    ws1['S'+str(row)] = float(FEB_IF_MON_NGENon)
    ws1['T'+str(row)] = float(BEB_IF_MON_NGENoff) # BEB channel A
    ws1['U'+str(row)] = float(BEB_IF_MON_NGENon) # BEB channel A
    ws1['V'+str(row)] = float(FEB_LD_MON_NGENoff)
    
    wb.save(str(folder+filename+'.xlsx'))
    
    print ("Continue file (y) or output file (n)?")
    ans=input("Please enter y or n:")
    if ((ans=='n') or (ans!='y')):
        break
    row=row+1
    
ws1['E1'] = str(t) # put date and time

print ("FEB Score Sheet data is saved")
wb.save(str(folder+filename+'.xlsx'))